from __future__ import annotations

from collections import defaultdict
from datetime import date
from io import BytesIO
from typing import Any, Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.config import get_settings
from backend.db.supabase_client import table_select_all
from backend.routers.common import DEMO_RETURNS, DEMO_SALES


DARK_BLUE = "1F3864"
RED = "FF0000"
LIGHT_BLUE = "BDD7EE"
WHITE = "FFFFFF"
LIGHT_GREY = "F2F2F2"
DARK_TEXT = "262626"
GREEN = "00B050"
ORANGE = "F4B183"
ACHIEVED_RED = "C00000"
DEFAULT_TARGET_VALUE = 50000000

CHANNEL_ORDER = ["AJIO_DROPSHIP", "AMAZON_IN_API", "MYNTRAPPMP", "MYNTRA SJIT", "NYKAA_FASHION", "TATACLIQ", "FLIPKART"]
CHANNEL_LABELS = {
    "MYNTRAPPMP": "MYNTRAPPMP",
    "MYNTRA": "MYNTRAPPMP",
    "MYNTRASJIT": "MYNTRA SJIT",
    "MYNTRA SJIT": "MYNTRA SJIT",
    "SJIT": "MYNTRA SJIT",
    "AJIO_DROPSHIP": "AJIO_DROPSHIP",
    "AJIO": "AJIO_DROPSHIP",
    "NYKAA_FASHION": "NYKAA_FASHION",
    "NYKAA": "NYKAA_FASHION",
    "FLIPKART": "FLIPKART",
    "AMAZON_IN_API": "AMAZON_IN_API",
    "AMAZON": "AMAZON_IN_API",
    "TATACLIQ": "TATACLIQ",
}
REPORT_COLUMNS = [
    "Date",
    "Platform",
    "Sale Qty",
    "Sale Value",
    "ASP",
    "Net Sale Qty",
    "Net Sale Value",
    "ASP",
    "RETURN Qty",
    "RETURN Value",
    "Return % Qty",
    "Return % Value",
    "Target",
]


def _month_start(day: date) -> date:
    return day.replace(day=1)


def _channel(value: Any) -> str:
    raw = str(value or "Unknown").strip()
    return CHANNEL_LABELS.get(raw.upper(), raw.upper())


def _amount(row: dict[str, Any]) -> float:
    return float(row.get("selling_price") or 0) * int(row.get("qty") or 0)


def _source(row: dict[str, Any]) -> str:
    return str(row.get("source") or "").strip().lower()


def _is_uniware_source(row: dict[str, Any]) -> bool:
    return _source(row) in {"unicommerce", "sales_master"}


def _select_dsr_sales_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    unicommerce_ppmp_dates = {
        str(row.get("date") or "")
        for row in rows
        if _source(row) == "unicommerce" and _channel(row.get("channel") or row.get("marketplace")) == "MYNTRAPPMP"
    }
    direct_ppmp_dates = {
        str(row.get("date") or "")
        for row in rows
        if _source(row) == "myntra_orders" and _channel(row.get("channel") or row.get("marketplace")) == "MYNTRAPPMP"
    }
    unicommerce_ppmp_dates.discard("")
    direct_ppmp_dates.discard("")
    if not unicommerce_ppmp_dates and not direct_ppmp_dates:
        return rows
    selected: list[dict[str, Any]] = []
    for row in rows:
        source = _source(row)
        channel = _channel(row.get("channel") or row.get("marketplace"))
        row_date = str(row.get("date") or "")
        if source == "myntra_orders" and channel == "MYNTRAPPMP" and row_date in unicommerce_ppmp_dates:
            continue
        if source == "sales_master" and channel == "MYNTRAPPMP" and row_date in direct_ppmp_dates:
            continue
        selected.append(row)
    return selected


def _safe_pct(numerator: float, denominator: float) -> float:
    return round((numerator * 100 / denominator), 2) if denominator else 0.0


def _safe_div(numerator: float, denominator: float) -> float:
    return round(numerator / denominator, 2) if denominator else 0.0


def _indian_number(value: float | int | None, decimals: int = 0) -> str:
    if value is None:
        return "-"
    sign = "-" if float(value) < 0 else ""
    absolute = abs(float(value))
    rounded = f"{absolute:.{decimals}f}" if decimals else str(int(round(absolute)))
    whole, _, fraction = rounded.partition(".")
    if len(whole) <= 3:
        grouped = whole
    else:
        grouped = whole[-3:]
        remaining = whole[:-3]
        parts = []
        while remaining:
            parts.insert(0, remaining[-2:])
            remaining = remaining[:-2]
        grouped = ",".join(parts + [grouped])
    return f"{sign}{grouped}{('.' + fraction) if decimals and fraction else ''}"


def _currency(value: float | int | None) -> str:
    return f"\u20b9{_indian_number(value)}" if value is not None else "-"


def _display(value: float | int | None, decimals: int = 0) -> str:
    return _indian_number(value, decimals=decimals)


def _percent_display(value: float | int | None) -> str:
    if value is None:
        return "-"
    return f"{round(float(value))}%"


def _dash_zero(value: float | int, decimals: int = 0) -> str:
    return "-" if not value else _display(value, decimals=decimals)


def _fetch_sales(start: date, end: date) -> list[dict[str, Any]]:
    filters: list[tuple[str, str]]
    if start == end:
        filters = [("date", f"eq.{start.isoformat()}")]
    else:
        filters = [("date", f"gte.{start.isoformat()}"), ("date", f"lte.{end.isoformat()}")]
    rows = table_select_all(
        "sales_fact",
        columns="date,channel,marketplace,source,selling_price,qty",
        filters=filters,
        max_rows=500000,
    )
    if not rows and not get_settings().has_supabase:
        rows = DEMO_SALES
    return _select_dsr_sales_rows(rows)


def _fetch_returns(start: date, end: date) -> list[dict[str, Any]]:
    filters: list[tuple[str, str]]
    if start == end:
        filters = [("date", f"eq.{start.isoformat()}")]
    else:
        filters = [("date", f"gte.{start.isoformat()}"), ("date", f"lte.{end.isoformat()}")]
    rows = table_select_all(
        "returns_fact",
        columns="date,channel,return_value,qty",
        filters=filters,
        max_rows=500000,
    )
    if not rows and not get_settings().has_supabase:
        rows = DEMO_RETURNS
    return rows


def _target_for_month(month_start: date) -> dict[str, Any] | None:
    try:
        rows = table_select_all(
            "targets",
            columns="id,month,channel,target_value,target_qty,created_by,created_at",
            filters=[("month", f"eq.{month_start.isoformat()}")],
            max_rows=1000,
        )
    except Exception as exc:
        text = str(exc).lower()
        if "targets" in text and ("pgrst205" in text or "could not find the table" in text or "404" in text):
            return {"id": None, "month": month_start.isoformat(), "channel": "ALL", "target_value": DEFAULT_TARGET_VALUE, "target_qty": 0, "setup_required": True}
        return None
    all_target = next((row for row in rows if str(row.get("channel") or "").upper() == "ALL"), None)
    return all_target or (rows[0] if rows else None)


def _aggregate(sales_rows: list[dict[str, Any]], return_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, float]] = defaultdict(lambda: {"sale_qty": 0.0, "sale_value": 0.0, "return_qty": 0.0, "return_value": 0.0})
    for row in sales_rows:
        channel = _channel(row.get("channel") or row.get("marketplace"))
        grouped[channel]["sale_qty"] += int(row.get("qty") or 0)
        grouped[channel]["sale_value"] += _amount(row)
    for row in return_rows:
        channel = _channel(row.get("channel"))
        grouped[channel]["return_qty"] += float(row.get("qty") or 0)
        grouped[channel]["return_value"] += float(row.get("return_value") or 0)

    channels = [channel for channel in CHANNEL_ORDER if channel in grouped]
    channels.extend(sorted(channel for channel in grouped if channel not in channels))
    rows = []
    for channel in channels:
        values = grouped[channel]
        sale_qty = int(values["sale_qty"])
        sale_value = round(values["sale_value"], 2)
        return_qty = int(values["return_qty"])
        return_value = round(values["return_value"], 2)
        net_qty = sale_qty - return_qty
        net_value = round(sale_value - return_value, 2)
        rows.append(
            {
                "platform": channel,
                "sale_qty": sale_qty,
                "sale_value": sale_value,
                "asp": _safe_div(sale_value, sale_qty),
                "net_sale_qty": net_qty,
                "net_sale_value": net_value,
                "net_asp": _safe_div(net_value, net_qty),
                "return_qty": return_qty,
                "return_value": return_value,
                "return_pct_qty": _safe_pct(return_qty, sale_qty),
                "return_pct_value": _safe_pct(return_value, sale_value),
            }
        )
    return rows


def _total(rows: list[dict[str, Any]]) -> dict[str, Any]:
    sale_qty = sum(int(row["sale_qty"]) for row in rows)
    sale_value = round(sum(float(row["sale_value"]) for row in rows), 2)
    return_qty = sum(int(row["return_qty"]) for row in rows)
    return_value = round(sum(float(row["return_value"]) for row in rows), 2)
    net_qty = sale_qty - return_qty
    net_value = round(sale_value - return_value, 2)
    return {
        "platform": "TOTAL",
        "sale_qty": sale_qty,
        "sale_value": sale_value,
        "asp": _safe_div(sale_value, sale_qty),
        "net_sale_qty": net_qty,
        "net_sale_value": net_value,
        "net_asp": _safe_div(net_value, net_qty),
        "return_qty": return_qty,
        "return_value": return_value,
        "return_pct_qty": _safe_pct(return_qty, sale_qty),
        "return_pct_value": _safe_pct(return_value, sale_value),
    }


def _achieved_status(pct: float | None) -> str:
    if pct is None:
        return "not_set"
    if pct >= 80:
        return "green"
    if pct >= 50:
        return "orange"
    return "red"


def build_dsr_report(report_date: date) -> dict[str, Any]:
    start = _month_start(report_date)
    daily_rows = _aggregate(_fetch_sales(report_date, report_date), _fetch_returns(report_date, report_date))
    mtd_rows = _aggregate(_fetch_sales(start, report_date), _fetch_returns(start, report_date))
    daily_total = _total(daily_rows)
    mtd_total = _total(mtd_rows)
    target = _target_for_month(start)
    target_value = int(target.get("target_value") or 0) if target else None
    achieved_pct = round(float(mtd_total["sale_value"]) * 100 / target_value, 2) if target_value else None
    return {
        "date": report_date.isoformat(),
        "month": start.isoformat(),
        "title": "NAYAM BY LAKSHITA",
        "target": {
            "id": target.get("id") if target else None,
            "month": target.get("month") if target else start.isoformat(),
            "channel": target.get("channel") if target else "ALL",
            "target_value": target_value,
            "target_qty": int(target.get("target_qty") or 0) if target else None,
            "achieved_pct": achieved_pct,
            "status": _achieved_status(achieved_pct),
        },
        "daily": {"label": "DSR", "summary": "Daily Sale & Return Summary", "rows": daily_rows, "total": daily_total},
        "mtd": {"label": "MTD", "summary": "Monthly Sale & Return Summary", "rows": mtd_rows, "total": mtd_total},
    }


def _section_table_rows(section: dict[str, Any], left_label: str, target: dict[str, Any], report_date: date) -> list[list[str]]:
    rows = []
    label_value = report_date.strftime("%d-%b-%y") if left_label == "Date" else report_date.strftime("%b").upper()
    for row in section["rows"]:
        rows.append(
            [
                label_value,
                row["platform"],
                _display(row["sale_qty"]),
                _display(row["sale_value"]),
                _display(row["asp"]),
                _display(row["net_sale_qty"]),
                _display(row["net_sale_value"]),
                _display(row["net_asp"]),
                _dash_zero(row["return_qty"]),
                _dash_zero(row["return_value"]),
                _percent_display(row["return_pct_qty"]),
                _percent_display(row["return_pct_value"]),
                "",
            ]
        )
    total = section["total"]
    rows.append(
        [
            "TOTAL",
            "",
            _display(total["sale_qty"]),
            _display(total["sale_value"]),
            _display(total["asp"]),
            _display(total["net_sale_qty"]),
            _display(total["net_sale_value"]),
            _display(total["net_asp"]),
            _dash_zero(total["return_qty"]),
            _dash_zero(total["return_value"]),
            _percent_display(total["return_pct_qty"]),
            _percent_display(total["return_pct_value"]),
            _currency(target["target_value"]) if target.get("target_value") else "\u2014",
        ]
    )
    achieved = f"Achieved {target['achieved_pct']:.0f}%" if target.get("achieved_pct") is not None else "Achieved \u2014"
    rows.append(["", "", "", "", "", "", "", "", "", "", "", "", achieved])
    return rows


def _set_range_style(worksheet, row: int, start_col: int, end_col: int, fill: str, font_color: str = WHITE, bold: bool = True) -> None:
    for col in range(start_col, end_col + 1):
        cell = worksheet.cell(row=row, column=col)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(color=font_color, bold=bold)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _achievement_fill(status: str | None) -> str:
    if status == "green":
        return GREEN
    if status == "orange":
        return ORANGE
    if status == "red":
        return ACHIEVED_RED
    return DARK_BLUE


def _write_section(worksheet, start_row: int, label: str, summary: str, left_header: str, rows: list[list[str]], target_status: str | None, mtd: bool = False) -> int:
    worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
    worksheet.cell(start_row, 1, label)
    worksheet.merge_cells(start_row=start_row, start_column=3, end_row=start_row, end_column=12)
    worksheet.cell(start_row, 3, summary)
    worksheet.cell(start_row, 13, "")
    _set_range_style(worksheet, start_row, 1, 2, LIGHT_BLUE if mtd else DARK_BLUE, DARK_TEXT if mtd else WHITE)
    _set_range_style(worksheet, start_row, 3, 12, LIGHT_BLUE if mtd else DARK_BLUE, DARK_TEXT if mtd else WHITE)
    _set_range_style(worksheet, start_row, 13, 13, LIGHT_BLUE if mtd else DARK_BLUE, DARK_TEXT if mtd else WHITE)
    worksheet.cell(start_row, 1).font = Font(size=18, bold=True, color=DARK_TEXT if mtd else WHITE)
    worksheet.cell(start_row, 3).font = Font(size=12, bold=True, color=DARK_TEXT if mtd else WHITE)

    header_row = start_row + 1
    headers = [left_header, *REPORT_COLUMNS[1:]]
    for index, header in enumerate(headers, start=1):
        worksheet.cell(header_row, index, header)
    _set_range_style(worksheet, header_row, 1, 12, RED)
    _set_range_style(worksheet, header_row, 13, 13, LIGHT_BLUE, DARK_TEXT)

    for row_offset, row_values in enumerate(rows, start=header_row + 1):
        is_total = row_values[0] == "TOTAL"
        is_achieved = str(row_values[-1]).startswith("Achieved")
        for column, value in enumerate(row_values, start=1):
            cell = worksheet.cell(row_offset, column, value)
            cell.alignment = Alignment(horizontal="center" if column != 2 else "left", vertical="center")
            if is_total:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)
                cell.font = Font(bold=True)
            elif is_achieved and column == 13:
                cell.fill = PatternFill("solid", fgColor=_achievement_fill(target_status))
                cell.font = Font(color=WHITE, bold=True)
            elif row_offset % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="FAFAFA")
    return header_row + len(rows) + 1


def generate_dsr_excel(report: dict[str, Any]) -> bytes:
    report_date = date.fromisoformat(report["date"])
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "DSR Report"
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    title = worksheet.cell(1, 1, report["title"])
    title.font = Font(size=20, bold=True)
    title.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 26

    daily_rows = _section_table_rows(report["daily"], "Date", report["target"], report_date)
    target_status = report["target"].get("status")
    next_row = _write_section(worksheet, 3, "DSR", report["daily"]["summary"], "Date", daily_rows, target_status, mtd=False)
    mtd_rows = _section_table_rows(report["mtd"], "Month", report["target"], report_date)
    _write_section(worksheet, next_row + 1, "MTD", report["mtd"]["summary"], "Month", mtd_rows, target_status, mtd=True)

    thin = Side(style="thin", color="000000")
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=13):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for column in range(1, 14):
        letter = get_column_letter(column)
        max_length = max(len(str(cell.value or "")) for cell in worksheet[letter])
        worksheet.column_dimensions[letter].width = min(max(max_length + 2, 9), 18)
    worksheet.column_dimensions["B"].width = 18
    worksheet.freeze_panes = "A4"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def generate_dsr_png(report: dict[str, Any]) -> bytes:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.patches import Rectangle

    report_date = date.fromisoformat(report["date"])
    daily_rows = _section_table_rows(report["daily"], "Date", report["target"], report_date)
    mtd_rows = _section_table_rows(report["mtd"], "Month", report["target"], report_date)

    column_widths = [1.35, 1.65, 1.05, 1.35, 0.8, 1.25, 1.35, 0.85, 1.1, 1.35, 1.2, 1.25, 1.35]
    x_positions = [0.0]
    for width in column_widths:
        x_positions.append(x_positions[-1] + width)
    total_width = x_positions[-1]
    title_height = 0.5
    section_height = 0.56
    header_height = 0.46
    row_height = 0.46
    gap_height = 0.28
    heights = [
        title_height,
        section_height,
        header_height,
        *([row_height] * len(daily_rows)),
        gap_height,
        section_height,
        header_height,
        *([row_height] * len(mtd_rows)),
    ]
    total_height = sum(heights)

    fig, ax = plt.subplots(figsize=(total_width, total_height), dpi=120)
    ax.axis("off")
    ax.set_xlim(0, total_width)
    ax.set_ylim(total_height, 0)
    fig.subplots_adjust(left=0, right=1, top=1, bottom=0)

    def mpl_color(value: str) -> str:
        return f"#{value}" if len(value) == 6 and all(char in "0123456789ABCDEFabcdef" for char in value) else value

    def draw_cell(
        y: float,
        start_col: int,
        end_col: int,
        height: float,
        text: str = "",
        fill: str = WHITE,
        font_color: str = "black",
        bold: bool = False,
        size: float = 10.0,
        align: Literal["left", "center", "right"] = "center",
    ) -> None:
        x = x_positions[start_col]
        width = x_positions[end_col + 1] - x
        ax.add_patch(Rectangle((x, y), width, height, facecolor=f"#{fill}", edgecolor="black", linewidth=0.65))
        if not text:
            return
        if align == "left":
            text_x = x + 0.06
        elif align == "right":
            text_x = x + width - 0.06
        else:
            text_x = x + width / 2
        ax.text(
            text_x,
            y + height / 2,
            text,
            ha=align,
            va="center",
            color=mpl_color(font_color),
            fontsize=size,
            fontweight="bold" if bold else "normal",
        )

    def draw_full_row(y: float, height: float, row: list[str], stripe: bool = False) -> None:
        is_total = row[0] == "TOTAL"
        is_achieved = str(row[-1]).startswith("Achieved")
        for col, value in enumerate(row):
            fill = LIGHT_GREY if is_total else "FAFAFA" if stripe else WHITE
            font_color = "black"
            bold = is_total
            if is_achieved and col == 12:
                fill = _achievement_fill(report["target"].get("status"))
                font_color = WHITE
                bold = True
            elif is_achieved:
                value = ""
            align: Literal["left", "center", "right"] = "right" if col >= 2 and not is_achieved else "center"
            if col == 1 and not is_total and not is_achieved:
                align = "left"
            draw_cell(y, col, col, height, value, fill, font_color, bold, 10.5, align)

    y = 0.0
    draw_cell(y, 0, 12, heights[0], report["title"], WHITE, "black", True, 16)
    y += heights[0]

    draw_cell(y, 0, 1, heights[1], "DSR", DARK_BLUE, WHITE, True, 15)
    draw_cell(y, 2, 12, heights[1], report["daily"]["summary"], DARK_BLUE, WHITE, True, 12)
    y += heights[1]

    for index, header in enumerate(["Date", *REPORT_COLUMNS[1:]]):
        draw_cell(y, index, index, heights[2], header, LIGHT_BLUE if index == 12 else RED, "black" if index == 12 else WHITE, True, 9.8)
    y += heights[2]

    for index, row in enumerate(daily_rows):
        draw_full_row(y, row_height, row, stripe=index % 2 == 1)
        y += row_height

    draw_cell(y, 0, 12, heights[3 + len(daily_rows)], "", WHITE)
    y += heights[3 + len(daily_rows)]

    draw_cell(y, 0, 1, section_height, "MTD", LIGHT_BLUE, "black", True, 15)
    draw_cell(y, 2, 12, section_height, report["mtd"]["summary"], LIGHT_BLUE, "black", True, 12)
    y += section_height

    for index, header in enumerate(["Month", *REPORT_COLUMNS[1:]]):
        draw_cell(y, index, index, header_height, header, LIGHT_BLUE if index == 12 else RED, "black" if index == 12 else WHITE, True, 9.8)
    y += header_height

    for index, row in enumerate(mtd_rows):
        draw_full_row(y, row_height, row, stripe=index % 2 == 1)
        y += row_height

    output = BytesIO()
    fig.savefig(output, format="png")
    plt.close(fig)
    return output.getvalue()
