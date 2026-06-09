from __future__ import annotations

from io import BytesIO
import unittest
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.main import app
from backend.routers import inventory


class InventoryReplenishmentRouteTest(unittest.TestCase):
    def test_category_status_matrix_includes_sales_and_return_rate(self) -> None:
        sku_rows = [
            {
                "style_color": "green-good",
                "category_new": "Green",
                "inventory_status": "INSTOCK",
                "total_inventory": 100,
                "ros_30d": 4,
            },
            {
                "style_color": "red-risk",
                "category_new": "Red",
                "inventory_status": "OOS",
                "total_inventory": 0,
                "ros_30d": 1,
            },
        ]

        def select_all(table: str, **kwargs):
            del kwargs
            if table == "sku_master":
                return sku_rows
            if table == "sales_fact":
                return [
                    {"date": "2026-05-03", "style_color": "green-good", "selling_price": 100, "qty": 10, "source": "myntra_orders"},
                    {"date": "2026-05-03", "style_color": "red-risk", "selling_price": 100, "qty": 5, "source": "myntra_orders"},
                    {"date": "2026-05-03", "style_color": "unmapped-style", "selling_price": 200, "qty": 1, "source": "myntra_orders"},
                ]
            if table == "returns_fact":
                return [
                    {"date": "2026-05-04", "style_color": "green-good", "return_value": 100, "qty": 1},
                    {"date": "2026-05-04", "style_color": "red-risk", "return_value": 500, "qty": 5},
                    {"date": "2026-05-04", "style_color": "unmapped-style", "return_value": 200, "qty": 1},
                ]
            return []

        with patch("backend.routers.inventory.table_select_all", side_effect=select_all):
            response = TestClient(app).get(
                "/api/inventory/category_status_matrix",
                params={"from_date": "2026-05-01", "to_date": "2026-05-31"},
            )

        self.assertEqual(response.status_code, 200)
        rows = {row["category"]: row for row in response.json()}
        self.assertEqual(rows["Green"]["sales_value"], 1000)
        self.assertEqual(rows["Green"]["sales_qty"], 10)
        self.assertEqual(rows["Green"]["return_pct"], 10)
        self.assertEqual(rows["Green"]["instock_sales_qty"], 10)
        self.assertEqual(rows["Green"]["instock_sales_mix_pct"], 100)
        self.assertEqual(rows["Green"]["instock_return_pct"], 10)
        self.assertEqual(rows["Green"]["broken_sales_qty"], 0)
        self.assertEqual(rows["Red"]["sales_value"], 500)
        self.assertEqual(rows["Red"]["return_pct"], 100)
        self.assertEqual(rows["Red"]["oos_sales_qty"], 5)
        self.assertEqual(rows["Red"]["oos_sales_mix_pct"], 100)
        self.assertEqual(rows["Red"]["oos_return_pct"], 100)
        self.assertEqual(rows["Unknown"]["sales_value"], 200)
        self.assertEqual(rows["Unknown"]["return_pct"], 100)
        self.assertEqual(rows["Grand Total"]["return_pct"], 43.75)
        self.assertEqual(rows["Grand Total"]["instock_sales_qty"], 10)
        self.assertEqual(rows["Grand Total"]["oos_sales_qty"], 5)

    def test_category_status_matrix_download_returns_matrix_and_style_detail_sheets(self) -> None:
        sku_rows = [
            {
                "style_color": "green-good",
                "category_new": "Green",
                "inventory_status": "INSTOCK",
                "total_inventory": 100,
                "ros_30d": 4,
                "doi": 25,
                "priority": "P1",
            },
            {
                "style_color": "red-risk",
                "category_new": "Red",
                "inventory_status": "OOS",
                "total_inventory": 0,
                "ros_30d": 1,
                "priority": "P0",
            },
        ]

        def select_all(table: str, **kwargs):
            del kwargs
            if table == "sku_master":
                return sku_rows
            if table == "sales_fact":
                return [
                    {"date": "2026-05-03", "style_color": "green-good", "selling_price": 100, "qty": 10, "source": "myntra_orders"},
                    {"date": "2026-05-03", "style_color": "red-risk", "selling_price": 100, "qty": 5, "source": "myntra_orders"},
                ]
            if table == "returns_fact":
                return [
                    {"date": "2026-05-04", "style_color": "green-good", "return_value": 100, "qty": 1},
                    {"date": "2026-05-04", "style_color": "red-risk", "return_value": 500, "qty": 5},
                ]
            return []

        with patch("backend.routers.inventory.table_select_all", side_effect=select_all):
            response = TestClient(app).get(
                "/api/inventory/category_status_matrix/download",
                params={"from_date": "2026-05-01", "to_date": "2026-05-31"},
            )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.headers["content-type"],
        )
        self.assertIn("category-status-matrix.xlsx", response.headers["content-disposition"])

        workbook = load_workbook(BytesIO(response.content), data_only=True)
        self.assertEqual(workbook.sheetnames, ["Category Matrix", "Style Details"])

        matrix_sheet = workbook["Category Matrix"]
        style_sheet = workbook["Style Details"]
        self.assertEqual([cell.value for cell in matrix_sheet[1]][:5], ["Category", "Sales Value", "Sale Qty", "Return Value", "Return Qty"])
        self.assertEqual([cell.value for cell in style_sheet[1]][:5], ["Style", "Category", "Status", "Total Inventory", "ROS 30D"])

        matrix_categories = {row[0].value for row in matrix_sheet.iter_rows(min_row=2)}
        self.assertIn("Green", matrix_categories)
        self.assertIn("Red", matrix_categories)
        self.assertIn("Grand Total", matrix_categories)

        style_rows = {row[0].value: [cell.value for cell in row] for row in style_sheet.iter_rows(min_row=2)}
        self.assertEqual(style_rows["green-good"][1], "Green")
        self.assertEqual(style_rows["green-good"][2], "INSTOCK")
        self.assertEqual(style_rows["green-good"][7], 1000)
        self.assertEqual(style_rows["green-good"][11], 10)
        self.assertEqual(style_rows["red-risk"][1], "Red")
        self.assertEqual(style_rows["red-risk"][2], "OOS")

    def test_category_styles_returns_style_level_metrics(self) -> None:
        sku_rows = [
            {
                "style_color": "green-good",
                "category_new": "Green",
                "inventory_status": "INSTOCK",
                "total_inventory": 100,
                "ros_30d": 4,
                "doi": 25,
                "priority": "P1",
            },
            {
                "style_color": "red-risk",
                "category_new": "Red",
                "inventory_status": "OOS",
                "total_inventory": 0,
                "ros_30d": 1,
            },
        ]

        def sku_select(table: str, **kwargs):
            del kwargs
            return sku_rows if table == "sku_master" else []

        def fact_select(table: str, **kwargs):
            del kwargs
            if table == "sku_master":
                return sku_rows
            if table == "sales_fact":
                return [{"date": "2026-05-03", "style_color": "green-good", "selling_price": 100, "qty": 10, "source": "myntra_orders"}]
            if table == "returns_fact":
                return [{"date": "2026-05-04", "style_color": "green-good", "return_value": 100, "qty": 1}]
            return []

        with patch("backend.routers.common.table_select_all", side_effect=sku_select), patch("backend.routers.inventory.table_select_all", side_effect=fact_select):
            response = TestClient(app).get(
                "/api/inventory/category_styles",
                params={"category": "Green", "from_date": "2026-05-01", "to_date": "2026-05-31"},
            )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["total"], 1)
        row = payload["items"][0]
        self.assertEqual(row["style_color"], "green-good")
        self.assertEqual(row["status"], "INSTOCK")
        self.assertEqual(row["sales_value"], 1000)
        self.assertEqual(row["sales_qty"], 10)
        self.assertEqual(row["return_qty"], 1)
        self.assertEqual(row["return_pct"], 10)

    def test_sku_rows_uses_paginated_reader_so_style_count_is_not_capped(self) -> None:
        full_rows = [{"style_color": f"nm{index:04d}-black"} for index in range(1001)]
        capped_rows = [{"style_color": "only-first-page"}]

        with (
            patch("backend.routers.inventory.table_select_all", return_value=full_rows, create=True) as paginated_reader,
            patch("backend.routers.inventory._safe_table_rows", return_value=capped_rows) as capped_reader,
        ):
            rows = inventory._sku_rows()

        self.assertEqual(len(rows), 1001)
        self.assertEqual(rows[0]["style_color"], "nm0000-black")
        paginated_reader.assert_called_once()
        capped_reader.assert_not_called()

    def test_safe_table_rows_uses_paginated_reader_for_bounded_fact_reads(self) -> None:
        expected_rows = [{"style_color": "nm001-red"}, {"style_color": "nm002-red"}]
        filters = [("snapshot_date", "gte.2026-05-01")]

        with patch("backend.routers.inventory.table_select_all", return_value=expected_rows) as paginated_reader:
            rows = inventory._safe_table_rows("inventory_fact", columns="style_color", filters=filters, max_rows=2500)

        self.assertEqual(rows, expected_rows)
        paginated_reader.assert_called_once_with(
            "inventory_fact",
            columns="style_color",
            filters=filters,
            max_rows=2500,
        )

    def test_safe_table_rows_returns_empty_list_when_paginated_reader_fails(self) -> None:
        with patch("backend.routers.inventory.table_select_all", side_effect=RuntimeError("temporary failure")):
            rows = inventory._safe_table_rows("inventory_fact", columns="style_color", max_rows=2500)

        self.assertEqual(rows, [])

    def test_replenishment_plan_considers_manual_uploads_and_size_split(self) -> None:
        sku_rows = [
            {
                "style_color": "nm001-red",
                "category_new": "NOOS",
                "sale_grade_old": "NOOS",
                "inventory_status": "INSTOCK",
                "total_inventory": 900,
                "ros_7d": 6,
                "ros_30d": 6,
                "ros": 6,
                "doi": 150,
            },
            {
                "style_color": "nm002-red",
                "category_new": "NOOS",
                "sale_grade_old": "NOOS",
                "inventory_status": "OOS",
                "total_inventory": 0,
                "ros_7d": 0,
                "ros_30d": 0,
                "ros": 0,
                "doi": 0,
            },
            {
                "style_color": "nm003-black",
                "category_new": "NOOS",
                "sale_grade_old": "NOOS",
                "inventory_status": "INSTOCK",
                "total_inventory": 300,
                "ros_7d": 10,
                "ros_30d": 8,
                "ros": 8,
                "doi": 34,
            },
        ]

        def rows_for(table: str, **_kwargs):
            if table == "replenishment_log":
                return [
                    {"style_color": "nm002-red", "replenishment_qty": 200, "status": "planned"},
                    {"style_color": "nm003-black", "replenishment_qty": 300, "status": "planned"},
                ]
            if table == "inventory_fact":
                return [
                    {"snapshot_date": "2026-05-24", "style_color": "nm002-red", "size": "S", "qty": 0},
                    {"snapshot_date": "2026-05-24", "style_color": "nm002-red", "size": "M", "qty": 0},
                    {"snapshot_date": "2026-05-24", "style_color": "nm002-red", "size": "L", "qty": 0},
                    {"snapshot_date": "2026-05-24", "style_color": "nm002-red", "size": "XL", "qty": 0},
                ]
            if table == "sales_fact":
                return [
                    {"date": "2026-05-24", "style_color": "nm002-red", "size": "S", "qty": 1},
                    {"date": "2026-05-24", "style_color": "nm002-red", "size": "M", "qty": 4},
                    {"date": "2026-05-24", "style_color": "nm002-red", "size": "L", "qty": 3},
                    {"date": "2026-05-24", "style_color": "nm002-red", "size": "XL", "qty": 2},
                    {"date": "2026-05-24", "style_color": "nm003-black", "size": "M", "qty": 10},
                ]
            if table == "visibility_fact":
                return []
            return []

        with (
            patch("backend.routers.inventory.table_select_all", return_value=sku_rows),
            patch("backend.routers.inventory._safe_table_rows", side_effect=rows_for),
        ):
            response = TestClient(app).get("/api/inventory/replenishment_plan")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["summary"]["already_planned_styles"], 2)
        self.assertEqual(payload["summary"]["manual_pending_qty"], 500)
        oos_item = next(item for item in payload["items"] if item["style_color"] == "nm002-red")
        covered_item = next(item for item in payload["items"] if item["style_color"] == "nm003-black")
        self.assertTrue(oos_item["already_planned"])
        self.assertEqual(oos_item["pending_replenishment_qty"], 200)
        self.assertEqual(oos_item["recommended_replenishment_qty"], 500)
        self.assertEqual(oos_item["size_replenishment"][0], {"size": "M", "current_qty": 0, "recommended_qty": 200})
        self.assertEqual(covered_item["urgency"], "Covered")
        self.assertEqual(covered_item["recommended_replenishment_qty"], 0)

    def test_filters_replenishment_history_by_status_and_date_range(self) -> None:
        rows = [
            {
                "id": 1,
                "style_color": "nm1292-navy",
                "replenishment_qty": 200,
                "replenishment_date": "2026-05-10",
                "uploaded_by": "ops@example.com",
                "notes": "planned for NOOS",
                "status": "planned",
                "created_at": "2026-05-10T09:00:00+00:00",
            },
            {
                "id": 2,
                "style_color": "nm1301-black",
                "replenishment_qty": 100,
                "replenishment_date": "2026-04-28",
                "uploaded_by": "ops@example.com",
                "notes": "outside range",
                "status": "planned",
                "created_at": "2026-04-28T09:00:00+00:00",
            },
            {
                "id": 3,
                "style_color": "23ssnm001-cherry",
                "replenishment_qty": 500,
                "replenishment_date": "2026-05-11",
                "uploaded_by": "ops@example.com",
                "notes": "completed style",
                "status": "completed",
                "created_at": "2026-05-11T09:00:00+00:00",
            },
        ]

        with patch("backend.routers.common.table_select_all", return_value=rows):
            response = TestClient(app).get(
                "/api/inventory/replenishment",
                params={
                    "status": "planned",
                    "from_date": "2026-05-01",
                    "to_date": "2026-05-31",
                },
            )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["total"], 1)
        self.assertEqual(payload["items"][0]["style_color"], "nm1292-navy")
        self.assertEqual(payload["items"][0]["status"], "planned")
