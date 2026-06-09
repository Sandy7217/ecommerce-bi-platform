from __future__ import annotations

import unittest
from io import BytesIO
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.main import app


def _rows_for(table: str, **_kwargs):
    if table == "sales_fact":
        return [
            {
                "date": "2026-05-02",
                "internal_sku": "SKU-1",
                "style_color": "nm001-red",
                "channel": "Myntra",
                "marketplace": "Myntra",
                "selling_price": 1000,
                "mrp": 1500,
                "discount": 500,
                "qty": 2,
                "order_id": "ORD-1",
                "state": "MH",
                "city": "Mumbai",
            }
        ]
    if table == "returns_fact":
        return [
            {
                "date": "2026-05-03",
                "internal_sku": "SKU-1",
                "style_color": "nm001-red",
                "channel": "Myntra",
                "qty": 1,
                "return_value": 1000,
                "return_type": "Customer Return",
                "state": "MH",
            }
        ]
    if table == "sku_master":
        return [
            {
                "style_color": "nm001-red",
                "category_new": "NOOS",
                "sale_grade_old": "NOOS",
                "inventory_status": "INSTOCK",
                "total_inventory": 4,
                "ros": 3.0,
                "ros_30d": 2.0,
                "doi": 2.0,
            }
        ]
    return []


class ReportsDownloadTest(unittest.TestCase):
    def test_download_sales_csv_with_expected_columns(self) -> None:
        with patch("backend.routers.reports.table_select_all", side_effect=_rows_for):
            response = TestClient(app).get(
                "/api/reports/download",
                params={"type": "sales", "format": "csv", "from_date": "2026-05-01", "to_date": "2026-05-31"},
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.headers["content-type"].split(";")[0], "text/csv")
        self.assertIn("attachment; filename=mtd-sales-report.csv", response.headers["content-disposition"])
        self.assertIn("Date,SKU,Style Color,Channel,Marketplace,Selling Price,MRP,Discount,Qty,Order ID,State,City", response.text)
        self.assertIn("2026-05-02,SKU-1,nm001-red,Myntra,Myntra,1000,1500,500,2,ORD-1,MH,Mumbai", response.text)

    def test_download_inventory_excel_with_expected_sheet(self) -> None:
        with patch("backend.routers.reports.table_select_all", side_effect=_rows_for):
            response = TestClient(app).get(
                "/api/reports/download",
                params={"type": "inventory", "format": "excel"},
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        sheet = workbook.active
        self.assertEqual(sheet.title, "Inventory Report")
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(headers, ["Style Color", "Category", "Sale Grade", "Inventory Status", "Total Inventory", "ROS 30D", "DOI", "Priority"])
        first_row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
        self.assertEqual(first_row[0], "nm001-red")
        self.assertTrue(str(first_row[-1]).startswith("P1"))

    def test_download_remaining_report_types_as_csv(self) -> None:
        cases = [
            ("replenishment", "Style Color,Category,Status,Total Inventory,ROS 30D,DOI,Priority,Recommended Qty"),
            ("returns", "Date,SKU,Style Color,Channel,Return Qty,Return Value,Return Type,State"),
            ("category_analysis", "Style Color,Old Sale Grade,New Category,Inventory Status,Total Inventory,ROS,MTD Sales"),
        ]
        for report_type, header in cases:
            with self.subTest(report_type=report_type):
                with patch("backend.routers.reports.table_select_all", side_effect=_rows_for):
                    response = TestClient(app).get(
                        "/api/reports/download",
                        params={"type": report_type, "format": "csv", "from_date": "2026-05-01", "to_date": "2026-05-31"},
                    )

                self.assertEqual(response.status_code, 200)
                self.assertIn(header, response.text)

    def test_empty_date_params_use_default_month_window(self) -> None:
        with patch("backend.routers.reports.table_select_all", side_effect=_rows_for):
            response = TestClient(app).get(
                "/api/reports/download",
                params={"type": "sales", "format": "csv", "from_date": "", "to_date": ""},
            )

        self.assertEqual(response.status_code, 200)
        self.assertIn("mtd-sales-report.csv", response.headers["content-disposition"])


if __name__ == "__main__":
    unittest.main()
